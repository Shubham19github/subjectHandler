import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def ptrn():
    print()
    print("*******************************************************************************")
    print("*******************************************************************************")
    print()

def create_records():
    counter=5
    wb =openpyxl.Workbook()
    wb.save(file_name)
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    sheet.title = sheet_name
    wb.create_sheet(index=1, title='Attendance_sheet')
    sheet1 = wb.get_sheet_by_name('Attendance_sheet')
    sheet.column_dimensions['A'].width = 30
    sheet1.column_dimensions['A'].width = 30
    sheet1.column_dimensions['B'].width = 15
    sheet1.column_dimensions['C'].width = 15
    sheet1.column_dimensions['D'].width = 15
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.freeze_panes = 'A5'
    sheet1.freeze_panes = 'A5'
    for i in range(2,12):
        sheet.column_dimensions[get_column_letter(i)].width = 15
    sheet.merge_cells('C1:G1')
    sheet1.merge_cells('C1:F1')
    sheet['C1']=sub_name+" ( "+sub_code+" )"
    sheet['C1'].font=fontObj2
    sheet1['C1'].alignment=align
    sheet1['C1']="ATTENDANCE SHEET ( "+sub_code+" ) - 2016"
    sheet1['C1'].font=fontObj2
    sheet['C1'].alignment=align
    sheet['A4']='NAME'
    sheet['A4'].font=fontObj1
    sheet['B4']='ROLL NO.'
    sheet['B4'].font=fontObj1
    sheet1['A4']='NAME'
    sheet1['A4'].font=fontObj1
    sheet1['B4']='ROLL NO.'
    sheet1['B4'].font=fontObj1
    sheet['C4']='TEST 1'
    sheet['C4'].font=fontObj1
    sheet['D4']='TEST 2'
    sheet['D4'].font=fontObj1
    sheet['E4']='MAJOR 1'
    sheet['E4'].font=fontObj1
    sheet['F4']='TEST 4'
    sheet['F4'].font=fontObj1
    sheet['G4']='TEST 5'
    sheet['G4'].font=fontObj1
    sheet['H4']='MAJOR 2'
    sheet['H4'].font=fontObj1
    sheet['I4']='TOTAL'
    sheet['I4'].font=fontObj1
    sheet['J4']='GRADE'
    sheet['J4'].font=fontObj1
    sheet['K4']='ATTENDANCE'
    sheet['K4'].font=fontObj1
    sheet1['C4']='TOTAL'
    sheet1['C4'].font=fontObj1
    sheet1['D4']='PERCENTAGE'
    sheet1['D4'].font=fontObj1
    sheet1['A2']='Total Class Taken  ->'
    sheet1['B2']=0

    f=open('student.txt','r')
    for line in f:
        a=str(counter)
        sheet['A'+a]=line[9:40]
        sheet['A'+a].font=fontObj3
        sheet['B'+a]=line[0:8]
        sheet['B'+a].font=fontObj3
        sheet['C'+a]=0
        sheet['D'+a]=0
        sheet['E'+a]=0
        sheet['F'+a]=0
        sheet['G'+a]=0
        sheet['H'+a]=0
        sheet['I'+a] = 0
        sheet['J'+a] = ' '
        sheet['K'+a] = 0
        sheet1['A'+a]=line[9:40]
        sheet1['A'+a].font=fontObj3
        sheet1['B'+a]=line[0:8]
        sheet1['B'+a].font=fontObj3
        sheet1['C'+a]=0
        sheet1['D'+a]=0
        counter=counter+1
    f.close()
    print("Records Entered Successfully")
    wb.save(file_name)
    
def view_records():
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name(sheet_name)
    sheet1 = wb.get_sheet_by_name('Attendance_sheet')
    total_class=sheet1.cell(row=2,column=column_index_from_string('B')).value
    outer_loop=True
    present=False
    while(outer_loop):
        roll=input("Enter Roll Number : ")
        for x in range(5, sheet.max_row+1):
            if(sheet.cell(row=x,column=column_index_from_string('B')).value)==roll:
                print("Name of the student : "+sheet.cell(row=x,column=column_index_from_string('A')).value)
                print("Test 1 : "+str(sheet.cell(row=x,column=column_index_from_string('C')).value)+"/25")
                print("Test 2 : "+str(sheet.cell(row=x,column=column_index_from_string('D')).value)+"/25")
                print("Major 1 : "+str(sheet.cell(row=x,column=column_index_from_string('E')).value)+"/40")
                print("Test 4 : "+str(sheet.cell(row=x,column=column_index_from_string('F')).value)+"/25")
                print("Test 5 : "+str(sheet.cell(row=x,column=column_index_from_string('G')).value)+"/25")
                print("Major 2 : "+str(sheet.cell(row=x,column=column_index_from_string('H')).value)+"/60")
                print("Total marks : "+str(sheet.cell(row=x,column=column_index_from_string('I')).value)+"/200")
                print("Grade Obtained : "+str(sheet.cell(row=x,column=column_index_from_string('J')).value))
                print("Attendance : "+str(sheet1.cell(row=x,column=column_index_from_string('C')).value)+"/"+str(total_class)+" = "+str(sheet1.cell(row=x,column=column_index_from_string('D')).value)+" %")
                print()
                present=True
                break
            else:
                continue
        if(present==False):
            print("Record not Found")
            print()
        opt=input("Continue ?? ( y/n ) : ")
        print()
        if(opt=='n'):
            outer_loop=False

def marks():
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name(sheet_name)
    outer_loop=True
    while(outer_loop):
        counter=5
        print(" 1 : Test 1")
        print(" 2 : Test 2")
        print(" 3 : Major 1")
        print(" 4 : Test 4")
        print(" 5 : Test 5")
        print(" 6 : Major 2")
        print(" 0 : None")
        print()
        option=input("Press : ")
        if(option=='0'):
            outer_loop=False
        elif(option=='1'):
            f=open('test1.txt','r')
            for line in f:
                a=str(counter)
                sheet['C'+a]=int(line)
                sheet.cell(row=counter,column=column_index_from_string('I')).value+=int(line)
                counter=counter+1
            f.close()
            wb.save(file_name)
            print("Test 1 marks updated.")
            ptrn()
   
        elif(option=='2'):
            f=open('test2.txt','r')
            for line in f:
                a=str(counter)
                sheet['D'+a]=int(line)
                sheet.cell(row=counter,column=column_index_from_string('I')).value+=int(line)
                counter=counter+1
            f.close()
            wb.save(file_name)
            print("Test 2 marks updated.")
            ptrn()
            
        elif(option=='3'):
            f=open('major1.txt','r')
            for line in f:
                a=str(counter)
                sheet['E'+a]=int(line)
                sheet.cell(row=counter,column=column_index_from_string('I')).value+=int(line)
                counter=counter+1
            f.close()
            wb.save(file_name)
            print("Major 1 marks updated.")
            ptrn()
        elif(option=='4'):
            f=open('test4.txt','r')
            for line in f:
                a=str(counter)
                sheet['F'+a]=int(line)
                sheet.cell(row=counter,column=column_index_from_string('I')).value+=int(line)
                counter=counter+1
            f.close()
            wb.save(file_name)
            print("Test 4 marks updated.")
            ptrn()
            
        elif(option=='5'):
            f=open('test5.txt','r')
            for line in f:
                a=str(counter)
                sheet['G'+a]=int(line)
                sheet.cell(row=counter,column=column_index_from_string('I')).value+=int(line)
                counter=counter+1
            f.close()
            wb.save(file_name)
            print("Test 5 marks updated.")
            ptrn()
            
        elif(option=='6'):
            f=open('major2.txt','r')
            for line in f:
                a=str(counter)
                sheet['H'+a]=int(line)
                sheet.cell(row=counter,column=column_index_from_string('I')).value+=int(line)
                counter=counter+1
            f.close()
            wb.save(file_name)
            print("Major 2 marks updated.")
            ptrn()
            
        else:
            print("Wrong Choice..Press Again")
            ptrn()
    
def grade():
    ap=150
    a=130
    bp=110
    b=90
    cp=80
    c=70
    d=60
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name(sheet_name)
    for x in range(5, sheet.max_row+1):
        if( ap <= sheet.cell(row=x,column=column_index_from_string('I')).value <= Total_marks):
                sheet.cell(row=x,column=column_index_from_string('J')).value='A+'
        elif( a <= sheet.cell(row=x,column=column_index_from_string('I')).value < ap):
                            sheet.cell(row=x,column=column_index_from_string('J')).value='A'
        elif( bp <= sheet.cell(row=x,column=column_index_from_string('I')).value < a):
                            sheet.cell(row=x,column=column_index_from_string('J')).value='B+'               
        elif( b <= sheet.cell(row=x,column=column_index_from_string('I')).value < bp):
                            sheet.cell(row=x,column=column_index_from_string('J')).value='B'
        elif( cp <= sheet.cell(row=x,column=column_index_from_string('I')).value < b):
                            sheet.cell(row=x,column=column_index_from_string('J')).value='C+'
        elif( c <= sheet.cell(row=x,column=column_index_from_string('I')).value < cp):
                            sheet.cell(row=x,column=column_index_from_string('J')).value='C'
        elif( d <= sheet.cell(row=x,column=column_index_from_string('I')).value < c):
                            sheet.cell(row=x,column=column_index_from_string('J')).value='D'
        else:
            sheet.cell(row=x,column=column_index_from_string('J')).value='F'

    print("Grades for every student updated")
    print()
    wb.save(file_name)

def attendance():
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name(sheet_name)
    sheet1 = wb.get_sheet_by_name('Attendance_sheet')
    total_class=sheet1.cell(row=2,column=column_index_from_string('B')).value
    name=["1.txt","2.txt","3.txt"]
    for i in name:
        counter=5
        k=sheet1.max_column+1
        f=open(i,'r')
        first=f.readline().rstrip()
        sheet1.cell(row=4,column=k).value=first
        total_class += 1
        sheet1.cell(row=2,column=column_index_from_string('B')).value+=1
        for line in f:
            a=str(counter)
            sheet1.cell(row=counter,column=k).value=int(line.rstrip())
            if(int(line.rstrip())==1):
                sheet1.cell(row=counter,column=column_index_from_string('C')).value += 1
            counter += 1

        f.close()
        
        for x in range(5, sheet.max_row+1):
            sheet1.cell(row=x,column=column_index_from_string('D')).value = sheet1.cell(row=x,column=column_index_from_string('C')).value / total_class *100
            sheet.cell(row=x,column=column_index_from_string('K')).value = sheet1.cell(row=x,column=column_index_from_string('D')).value
            
    print("Attendance updated.")
    wb.save(file_name)
    print()

def save():
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name(sheet_name)
    sheet1 = wb.get_sheet_by_name('Attendance_sheet')
    for row in sheet1.rows:
        for cell in row:
            cell.alignment=align
            
    for row in sheet.rows:
        for cell in row:
            cell.alignment=align
            
    print("Changes Saved.")
    wb.save(file_name)
    print()
    
Total_marks=200
no_of_students=0
counter=0
file_name='students.xlsx'
sheet_name='Students_marks'
fontObj1 = Font(name='Times New Roman', bold=True,color='323232',underline="single")
fontObj2 = Font(name='Times New Roman', bold=True,color='FC441C',underline="single",size=12)
fontObj3 = Font(name='Times New Roman',color='06DB80')
align=Alignment(horizontal='center')
ptrn()
print("Welcome to Subject Handler of Tezpur University.")
ptrn()
sub_name='PRINCIPLES OF PROGRAMMING LANGUAGES'
sub_code='CO304'
print("SUBJECT NAME : "+sub_name)
print()
print("COURSE CODE : "+sub_code)
ptrn()
while(1):
    print(" r : create record.")
    print(" m : update marks.")
    print(" g : grades.")
    print(" v : view records.")
    print(" a : attendance.")
    print(" s : save")
    print(" q : quit.")
    print()
    option=input("Press : ")
    ptrn()
    print()
    if(option=='r'):
        create_records()
        ptrn()
    elif(option=='m'):
        marks()
        ptrn()
    elif(option=='g'):
        grade()
        ptrn()
    elif(option=='v'):
        view_records()
        ptrn()
    elif(option=='a'):
        attendance()
        ptrn()
    elif(option=='s'):
        save()
        ptrn()
    elif(option=='q'):
        print()
        print("*********************** Thanks for using this handler *************************")
        print()
        exit()
    else:
        print("Wrong Choice..Press Again.")
        print()
